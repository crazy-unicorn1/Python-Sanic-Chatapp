# -*- coding: UTF-8 -*-
import datetime, random, string, re, os
from openpyxl import Workbook
from constants import SITE_CONFIG_CACHE, LANGUAGE, PermissionCls, ExportStatu, OnlineStatu, SITE_DICT_CACHE, SERVICE_CONNECTION, OPERATION_TYPES
from models.kefu_table import BlacklistTable, CustomerTable, ChatConversationTable, LeavingMessageTable, QuickReplyTable, problemTable, systemLogTable, ChatContentTable, FinishListTable
from models.site_table import SiteTable, ExportDataModel
from models.cms_user import CmsUserModel
from common_utils import xtjson
from modules.api_module.chat_tools import xtjsonCls
from common_utils.utils_funcs import update_language
from modules.api_module.chat_tools import emit

# 获取网站Code
def get_site_code():
    ''' 获取网站Code '''
    while True:
        _code = 'chat_'
        for i in range(6):
            _code += random.choice(string.ascii_letters + string.digits)
        if SiteTable.find_one({'site_code': _code}):
            continue
        return _code


# 获取域名列表
def getSiteDomain(MAIN_DOMAIN):
    domain_ls = [
        MAIN_DOMAIN,
    ]
    site_domain = ''
    if hasattr(SITE_CONFIG_CACHE, 'site_domain'):
        site_domain = SITE_CONFIG_CACHE.site_domain or ''
    for doamin in site_domain.split('\n'):
        doamin = doamin.strip()
        if not doamin:
            continue
        if doamin.count('.') > 2 or 'http' in doamin:
            continue
        if not doamin.replace('.', '').isalnum():
            continue
        if doamin not in domain_ls:
            domain_ls.append(doamin)
    return domain_ls


# 添加用户html
def add_user_html(current_admin_dict, crr_language, MAIN_DOMAIN):
    crr_role_code = current_admin_dict.get('role_code')
    html = ''
    html += '''
    <div class="addUserBox">
        <div style="height: 28rem; position: relative; box-sizing: border-box; overflow-y: auto;">
            <div class="portrait">
                <span style="float: left; width: 150px; text-align: right; line-height: 60px; font-size: 12px;">头像</span>
                <div class="img">
                    <img class="portrait_img" src="/assets/chat/images/defhead.png" alt="">
                </div>

                <div class="file-button" style="width: 80px;border-radius: 3px;border: solid 1px #eeeeee; padding: 6px 12px; font-size: 12px; display: inline-block; float: left; position: relative; top: 13px; cursor: pointer;">
                    <span>点击上传</span>
                    <input type="hidden" id="portrait" value="" aria-label="">
                    <input type="file" id="upload1" onchange="upload_file_func($(\'#upload1\'),$(\'#portrait\'),\'updatePortrait\', \'/site_admin/userManage\', $(\'.portrait_img\'), \'\', \'progress\')">
                </div>

                <div style="clear: both;width: 100%;overflow: hidden;height: 10px;"></div>
                <p style="margin-bottom: 0;font-size: 12px;text-align: left;padding-left: 160px;">仅限png格式；尺寸：144px X 144px；大小：不超过1M</p>
            </div>

            <div id="showbar" style="display: none"><div class="list-group-item"><script>dumprogressbar()</script></div></div>

            <div class="list-group-item">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>账户：</span>
                <input type="text" class="form-control" id="account" placeholder="账户" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
            </div>          
    '''

    if crr_role_code == PermissionCls.SUPERADMIN:
        html += f'''
            <div class="list-group-item">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>角色：</span>
                <select class="form-control" id="role_code" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="checkRole_func($(this).val())">
                    <option value="">选择角色</option>
                    <option value="agentadmin">代理商</option>
                    <option value="administrator">管理员</option>
                    <option value="customerservice">客服</option>
                </select>
            </div>                                                                           
        '''
    elif crr_role_code == PermissionCls.AgentAdmin:
        html += f'''
            <div class="list-group-item">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>角色：</span>
                <select class="form-control" id="role_code" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="checkRole_func($(this).val())">
                    <option value="">选择角色</option>
                    <option value="administrator">管理员</option>
                    <option value="customerservice">客服</option>
                </select>
            </div>                                                                           
        '''
    else:
        html += f'''
            <div class="list-group-item">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>角色：</span>
                <select class="form-control" id="role_code" aria-label="" style="display: inline-block; width: calc(100% - 180px)" onchange="checkRole_func($(this).val())">
                    <option value="">选择角色</option>
                    <option value="customerservice">客服</option>
                </select>
            </div>
        '''

    html += '''
        <div class="list-group-item">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>密码：</span>
            <input type="text" class="form-control" id="password" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>
        <div class="list-group-item">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>确认密码：</span>
            <input type="text" class="form-control" id="confirmPassword" placeholder="确认密码" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>
        <div class="list-group-item">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>姓名：</span>
            <input type="text" class="form-control" id="username" placeholder="姓名" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>            
    '''

    if crr_role_code in [PermissionCls.AgentAdmin, PermissionCls.SUPERADMIN]:
        html += '''
            <div class="list-group-item" style="display: none;">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>负责网站：</span>
                <select class="form-control" id="responsible_site" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    <option value="">选择负责网站</option>                 
        '''
        site_datas = [{}]
        if current_admin_dict.get('role_code') == PermissionCls.SUPERADMIN:
            site_datas = SiteTable.find_many({})
        if current_admin_dict.get('role_code') == PermissionCls.AgentAdmin:
            _cdas = CmsUserModel.find_many({'super_admin_id': current_admin_dict.get('uuid')})
            site_dis = []
            for _cd in _cdas:
                site_dis.append(_cd.get('responsible_site'))
            site_datas = SiteTable.find_many({'site_code': {'$in': site_dis}})

        for sd in site_datas:
            html += f'''
                <option value="{sd.get('site_code')}" >{sd.get('site_name')}</option>
            '''
        html += '''
                    </select>
                </div>        
        '''

    html += '''                                    
        <div class="list-group-item">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;">手机号：</span>
            <input type="text" class="form-control" id="telephone" placeholder="手机号" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>
        <div class="list-group-item">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;">昵称：</span>
            <input type="text" class="form-control" id="nickname" placeholder="昵称" value="" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>
        <div class="list-group-item">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;">邮箱：</span>
            <input type="text" class="form-control" id="email" placeholder="邮箱" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>
    '''

    html += '''
        <div class="list-group-item" style="display: none;">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>接入网站上限：</span>
            <input type="number" class="form-control" id="create_admin_count" placeholder="接入网站上限" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>        
        <div class="list-group-item" style="display: none;">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>创建网站客服上限：</span>
            <input type="number" class="form-control" id="create_cust_service_count" placeholder="创建网站客服上限" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>        
    '''

    html += f'''
        <div class="list-group-item" style="display: none;">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>接待上限：</span>
            <input type="number" class="form-control" id="reception_count" placeholder="接待上限" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>           
        <div class="list-group-item" style="display: none;">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>租用到期时间：</span>                    
            <input type="text" class="form-control selectDateYMD" id="zy_finish_time" onmouseenter="ddd()" value="" placeholder="代理商租用到期时间" aria-describedby="emailHelp" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>
        <div class="list-group-item" style="display: none;">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>网站到期时间：</span>                    
            <input type="text" class="form-control selectDateYMD" id="finish_time" onmouseenter="ddd()" value="" placeholder="网站到期时间" aria-describedby="emailHelp" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>
        <div class="list-group-item" style="display: none;">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>网站名称：</span>
            <input type="text" class="form-control" id="site_name" value="" placeholder="网站名称" aria-describedby="emailHelp" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>
        <div class="list-group-item" style="display: none;">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>网站链接：</span>
            <input type="text" class="form-control" id="site_link" value="" placeholder="网站链接" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
        </div>
    '''
    html += f'''
        <div class="list-group-item" style="display: none;">
            <span style="width: 120px; text-align: right; display: inline-block; position: relative;">聊天界面语言：</span>
            <select name="" class="form-control" id="site_language" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                <option value="">选择聊天界面语言</option>        
    '''
    for l in LANGUAGE.name_arr:
        html += f'''
        <option value="{l}">{LANGUAGE.name_dict.get(l)}</option>
        '''
    html += '''
            </select>
        </div>
    '''

    html += '''
            <div class="list-group-item" style="display: none;">
                <span style="width: 120px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>选择客服访问域名：</span>
                <select name="" class="form-control" id="use_domain" aria-label="" style="display: inline-block; width: calc(100% - 180px)">
                    <option value="">选择客服访问域名</option>
    '''
    domain_ls = getSiteDomain(MAIN_DOMAIN)
    for domain in domain_ls:
        html += f'''
        <option value="{domain}">{domain}</option>
        '''
    html += '''
                </select>
            </div>
        '''

    html += '''
        </div>

        <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

        <div style="position: relative; text-align: center">
            <span class="kfConfirmBtn" onclick="add_user_data()">确定</span>
            <span class="kfCancelBtn" onclick="xtalert.close()">取消</span>
        </div>

    </div>
    '''
    html = update_language(crr_language, html)
    return html


# 添加用户
def add_user_data(request_data, current_admin_dict, MAIN_DOMAIN):
    data_from = {}
    crr_role_code = current_admin_dict.get('role_code')

    portrait = request_data.get('portrait')
    account = request_data.get('account')
    telephone = request_data.get('telephone')
    password = request_data.get('password')
    username = request_data.get('username')
    nickname = request_data.get('nickname')
    role_code = request_data.get('role_code')
    reception_count = request_data.get('reception_count')
    responsible_site = request_data.get('responsible_site')
    email = request_data.get('email')
    create_admin_count = request_data.get('create_admin_count')
    finish_time = request_data.get('finish_time')
    zy_finish_time = request_data.get('zy_finish_time')
    site_name = request_data.get('site_name')
    site_link = request_data.get('site_link') or ''
    site_language = request_data.get('site_language') or LANGUAGE.en_US
    use_domain = request_data.get('use_domain') or ''
    create_cust_service_count = request_data.get('create_cust_service_count') or ''

    if not account or not password or not username or not role_code or not account or not password.strip() or not role_code:
        return xtjson.json_params_error('缺少数据！')

    if len(password) < 5 or len(password) > 18:
        return xtjson.json_params_error('密码要在5~18位数之间！')

    _reception_count = 0
    if reception_count is not None and str(reception_count).isdigit():
        _reception_count = int(reception_count)

    if CmsUserModel.find_one({'account': account}):
        return xtjson.json_params_error('账户已存在！')

    data_from['portrait'] = portrait or ''
    data_from['account'] = account
    data_from['telephone'] = telephone
    data_from['username'] = username
    data_from['nickname'] = nickname
    data_from['role_code'] = role_code
    data_from['email'] = email
    data_from['online_statu'] = OnlineStatu.online
    data_from['password'] = CmsUserModel.encry_password(password.strip())
    data_from['super_admin_id'] = current_admin_dict.get('uuid')
    data_from['is_activate'] = False
    data_from['statu'] = True
    data_from['language'] = LANGUAGE.zh_CN

    if role_code in [PermissionCls.Administrator, PermissionCls.CustomerService]:
        data_from['reception_count'] = _reception_count or 10

    if role_code and role_code == 'customerservice':
        if crr_role_code in [PermissionCls.SUPERADMIN, PermissionCls.AgentAdmin]:
            if not responsible_site:
                return xtjson.json_params_error('请选择负责网站！')
            data_from['responsible_site'] = responsible_site

            _crr_site_data = SITE_DICT_CACHE.get(responsible_site) or {}
            _create_cust_service_count = _crr_site_data.get('create_cust_service_count') or 0
            if _create_cust_service_count <= CmsUserModel.count({'responsible_site': responsible_site, 'role_code': PermissionCls.CustomerService}):
                return xtjson.json_params_error('创建客服数量已达上限！')

        elif crr_role_code == PermissionCls.Administrator:
            data_from['responsible_site'] = current_admin_dict.get('responsible_site')
            responsible_site = current_admin_dict.get('responsible_site')
            data_from['responsible_site'] = responsible_site

            _crr_site_data = SITE_DICT_CACHE.get(responsible_site) or {}
            __create_cust_service_count = _crr_site_data.get('create_cust_service_count') or 0
            crr_c_count = CmsUserModel.count({'responsible_site': responsible_site, 'role_code': PermissionCls.CustomerService}) or 0
            if __create_cust_service_count <= crr_c_count:
                return xtjson.json_params_error('创建客服数量已达上限！')

    crr_site_code = ''
    add_site_data_from = {}
    if role_code and role_code == 'administrator':
        if not finish_time or not site_name or not site_name.strip() or not site_link or not site_link.strip():
            return xtjson.json_params_error('缺少数据！')

        if current_admin_dict.get('role_code') == PermissionCls.AgentAdmin:
            _ccc = 0
            for d in CmsUserModel.find_many({'super_admin_id': current_admin_dict.get('uuid'), 'role_code': PermissionCls.Administrator}):
                _sddata = SiteTable.find_one({'uuid': d.get('responsible_site')}) or {}
                if _sddata:
                    _c_create_cust_service_count = _sddata.get('create_cust_service_count') or 0
                    _ccc += _c_create_cust_service_count

            _low_create_cust_service_count = int(current_admin_dict.get('create_cust_service_count') or 0)
            if int(create_cust_service_count) > _low_create_cust_service_count - _ccc:
                return xtjson.json_params_error('当前分配客服余量不足！')

        domain_ls = getSiteDomain(MAIN_DOMAIN)
        crr_site_code = get_site_code()
        add_site_data_from['link'] = site_link
        add_site_data_from['site_name'] = site_name
        add_site_data_from['use_domain'] = use_domain or random.choice(domain_ls)
        add_site_data_from['site_language'] = site_language or LANGUAGE.zh_CN
        add_site_data_from['site_code'] = crr_site_code
        add_site_data_from['site_google_verify_statu'] = False
        add_site_data_from['fast_state'] = False
        add_site_data_from['beep_switch'] = True
        add_site_data_from['finish_time'] = datetime.datetime.strptime(finish_time, '%Y-%m-%d')
        add_site_data_from['create_cust_service_count'] = int(create_cust_service_count or 0)
        data_from['responsible_site'] = crr_site_code

    if role_code and role_code == 'agentadmin':
        if not zy_finish_time:
            return xtjson.json_params_error('请选择到期时间！')
        data_from['zy_finish_time'] = datetime.datetime.strptime(zy_finish_time, '%Y-%m-%d')
        if not create_admin_count:
            return xtjson.json_params_error('请输入接入网站上限！')
        data_from['create_admin_count'] = int(create_admin_count or 0)
        data_from['create_cust_service_count'] = int(create_cust_service_count or 0)

    if add_site_data_from and crr_site_code:
        SiteTable.insert_one(add_site_data_from)
        while SITE_DICT_CACHE:
            for k in list(SITE_DICT_CACHE):
                SITE_DICT_CACHE.pop(k)
        _sds = SiteTable.find_many({}) or []
        for sd in _sds:
            SITE_DICT_CACHE[sd.get('site_code')] = sd

    CmsUserModel.insert_one(data_from)
    return


# 编辑用户html
def edit_user_html(data_uuid, language):
    _user_data = CmsUserModel.find_one({'uuid': data_uuid})
    if not _user_data:
        return xtjson.json_params_error('数据不存在！')
    html = f'''
    <div class="addUserBox">
        <div style="height: 23rem; position: relative; box-sizing: border-box; overflow-y: auto;">
            <div class="portrait">
                <span style="float: left; width: 55px; text-align: right; line-height: 60px; font-size: 12px;">头像</span>
                <div class="img">
                    <img class="portrait_img" src="{_user_data.get('portrait') or '/assets/chat/images/defhead.png'}" alt="">
                </div>
                <div class="file-button" style="width: 80px;border-radius: 3px;border: solid 1px #eeeeee; padding: 6px 12px; font-size: 12px; display: inline-block; float: left; position: relative; top: 13px; cursor: pointer;">
                    <span>点击上传</span>
                    <input type="hidden" id="portrait_edit" value="{_user_data.get('portrait') or ''}" aria-label="">
                    <input type="file" id="upload1" onchange="upload_file_func($(\'#upload1\'),$(\'#portrait_edit\'),\'updatePortrait\', \'/site_admin/userManage\', $(\'.portrait_img\'), \'\', \'progress\')">                        
                </div>

                <div style="clear: both;width: 100%;overflow: hidden;height: 10px;"></div>                    
                <p style="margin-bottom: 0;font-size: 12px;text-align: left;padding-left: 100px;">仅限png格式；尺寸：144px X 144px；大小：不超过1M</p>                    
            </div>                
    '''

    html += f'''
    <div class="list-group-item">
        <span style="width: 50px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>账户：</span>
        <input type="text" class="form-control" id="account_edit" placeholder="账户" value="{_user_data.get('account') or ''}" aria-label="" style="display: inline-block; width: calc(100% - 70px)" disabled>
    </div>
    '''

    html += f'''
    <div class="list-group-item">
        <span style="width: 50px; text-align: right; display: inline-block; position: relative;">新密码：</span>
        <input type="text" class="form-control" id="newPassword_edit" placeholder="密码" value="" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
    </div>
    '''
    html += '''
    <div class="list-group-item">
        <span style="width: 50px; text-align: right; display: inline-block; position: relative;">确认密码：</span>
        <input type="text" class="form-control" id="confirmPassword_edit" placeholder="密码" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
    </div>
    '''

    html += f'''
    <div class="list-group-item">
        <span style="width: 50px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>姓名：</span>
        <input type="text" class="form-control" id="username_edit" placeholder="姓名" value="{_user_data.get('username') or ''}" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
    </div>                         
    <div class="list-group-item">
        <span style="width: 50px; text-align: right; display: inline-block; position: relative;">手机号：</span>
        <input type="text" class="form-control" id="telephone_edit" placeholder="手机号" value="{_user_data.get('telephone') or ''}" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
    </div>        
    '''
    html += f'''
    <div class="list-group-item">
        <span style="width: 50px; text-align: right; display: inline-block; position: relative;">昵称：</span>
        <input type="text" class="form-control" id="nickname_edit" placeholder="昵称" value="{_user_data.get('nickname') or ''}" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
    </div>
    '''
    html += f'''
    <div class="list-group-item">
        <span style="width: 50px; text-align: right; display: inline-block; position: relative;">邮箱：</span>
        <input type="text" class="form-control" id="email_edit" placeholder="邮箱" value="{_user_data.get('email') or ''}" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
    </div>
    '''

    crr_role_code = _user_data.get('role_code')
    if crr_role_code == 'customerservice' or crr_role_code == 'administrator':
        html += f'''
        <div class="list-group-item">
            <span style="width: 50px; text-align: right; display: inline-block; position: relative;">接待上限：</span>
            <input type="number" class="form-control" id="reception_count" value="{_user_data.get('reception_count') or 0}" placeholder="接待上限" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
        </div>
        '''

    if crr_role_code == PermissionCls.AgentAdmin:
        html += f'''
            <div class="list-group-item">
                <span style="width: 50px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>租用到期时间：</span>
                <input type="text" class="form-control selectDateYMD" id="zy_finish_time" onmouseenter="ddd()" value="{ _user_data.get('zy_finish_time').strftime('%Y-%m-%d') if _user_data.get('zy_finish_time') else '' }" placeholder="代理商租用到期时间" aria-describedby="emailHelp" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
            </div>
        '''
        html += f'''
            <div class="list-group-item">
                <span style="width: 50px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>接入网站上限：</span>
                <input type="number" class="form-control" id="create_admin_count" value="{_user_data.get('create_admin_count')}" placeholder="接入网站上限" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
            </div>
            <div class="list-group-item">
                <span style="width: 50px; text-align: right; display: inline-block; position: relative;"><span class="text-danger">*</span>创建客服上限：</span>
                <input type="number" class="form-control" id="create_cust_service_count" value="{ _user_data.get('create_cust_service_count') or 0 }" placeholder="创建客服上限" aria-label="" style="display: inline-block; width: calc(100% - 70px)">
            </div>            
        '''

    html += '''
        </div>
    '''

    html += f'''
            <div class="blank" style="background: #eeeeee; height: 1px; margin: 15px 0;"></div>

            <div style="position: relative; text-align: center">
                <span class="kfConfirmBtn" onclick="edit_user_data('{data_uuid}', '{ _user_data.get('role_code') or '' }')">确定</span>
                <span class="kfCancelBtn" onclick="xtalert.close()">取消</span>
            </div>
    </div>
    '''
    html = update_language(language, html)
    return html


# 编辑用户信息
def edit_user_data(data_uuid, request_data):
    if not data_uuid:
        return xtjson.json_params_error('缺少数据id！')
    _data = CmsUserModel.find_one({'uuid': data_uuid})
    if not _data:
        return xtjson.json_params_error('数据不存在！')

    data_from = {}
    telephone = request_data.get('telephone') or ''
    username = request_data.get('username') or ''
    nickname = request_data.get('nickname') or ''
    portrait = request_data.get('portrait') or ''
    email = request_data.get('email') or ''
    newPassword = request_data.get('newPassword') or ''
    reception_count = request_data.get('reception_count') or ''
    create_admin_count = request_data.get('create_admin_count') or 0
    create_cust_service_count = request_data.get('create_cust_service_count') or 0
    data_from['telephone'] = telephone.strip()
    data_from['username'] = username.strip()
    data_from['nickname'] = nickname.strip()
    data_from['email'] = email.strip()
    data_from['portrait'] = portrait.strip()

    _reception_count = 0
    if reception_count is not None and str(reception_count).isdigit():
        _reception_count = int(reception_count)

    crr_role_code = _data.get('role_code')
    if crr_role_code in [PermissionCls.Administrator, PermissionCls.CustomerService]:
        data_from['reception_count'] = _reception_count or 10

    if newPassword:
        if len(newPassword) < 6 or len(newPassword) > 16:
            return xtjson.json_params_error('密码要在6~16位数之间！')
        data_from['password'] = CmsUserModel.encry_password(newPassword.strip())

    if crr_role_code and crr_role_code == 'agentadmin':
        zy_finish_time = request_data.get('zy_finish_time')
        if not zy_finish_time:
            return xtjson.json_params_error('请选择到期时间！')

        data_from['zy_finish_time'] = datetime.datetime.strptime(zy_finish_time, '%Y-%m-%d')
        data_from['create_admin_count'] = create_admin_count or 0
        data_from['create_cust_service_count'] = int(create_cust_service_count or 0)

    _data.update(data_from)
    CmsUserModel.save(_data)
    return


# 编辑网站信息
def site_form_html(language, data_uuid, MAIN_DOMAIN):
    crr_data = {}
    errormsg1 = '数据不存在！'
    errormsg1 = update_language(language, errormsg1)

    if data_uuid:
        crr_data = SiteTable.find_one({'uuid': data_uuid}) or {}
        if not crr_data:
            return xtjson.json_params_error(errormsg1)

    html = f'''        
            <div class="form-group">
                <label for="site_name" class="loglable" style="font-size: 13px; color: rgba(4,15,66,.6);">网站名称</label>
                <input type="text" class="form-control" id="site_name" value="{crr_data.get('site_name') or ''}" placeholder="网站名称" aria-describedby="emailHelp" aria-label="">
            </div>
            <div class="form-group">
                <label for="site_name" class="loglable" style="font-size: 13px; color: rgba(4,15,66,.6);">客服上限</label>
                <input type="text" class="form-control" id="create_cust_service_count" value="{crr_data.get('create_cust_service_count') or 0}" placeholder="客服上限" aria-describedby="emailHelp" aria-label="">
            </div>
            <div class="form-group">
                <label for="link" class="loglable" style="font-size: 13px; color: rgba(4,15,66,.6);">网站链接</label>
                <input type="text" class="form-control" id="link" value="{crr_data.get('link') or ''}" id="link" placeholder="网站链接" aria-label="">
            </div>              
            <div class="form-group">
                <label for="link" class="loglable" style="font-size: 13px; color: rgba(4,15,66,.6);">聊天界面语言</label>
                <select name="" class="form-control" id="site_language" aria-label="">
                    <option value="">选择聊天界面语言</option>
    '''
    for l in LANGUAGE.name_arr:
        html += f'''
        <option value="{l}" {'selected' if crr_data.get('site_language') == l else ''}>{LANGUAGE.name_dict.get(l)}</option>
        '''
    html += '''
                </select>                        
            </div>
        '''
    html += '''
            <div class="form-group">
                <label for="link" class="loglable" style="font-size: 13px; color: rgba(4,15,66,.6);">选择客服访问域名</label>
                <select name="" class="form-control" id="use_domain" aria-label="">
                    <option value="">选择客服访问域名</option>
    '''
    domain_ls = getSiteDomain(MAIN_DOMAIN)
    for domain in domain_ls:
        html += f'''
        <option value="{domain}" {'selected' if crr_data.get('use_domain') == domain else ''}>{domain}</option>
        '''
    html += '''
                </select>
            </div>
        '''

    html += f'''    
            <div class="form-group">
                <label for="link" class="loglable" style="font-size: 13px; color: rgba(4,15,66,.6);">网站到期时间</label>
                <input type="text" class="form-control selectDateYMD" id="finish_time" onmouseenter="ddd()" value="{ crr_data.get('finish_time').strftime('%Y-%m-%d') if crr_data.get('finish_time') else '' }" placeholder="网站到期时间" aria-describedby="emailHelp" aria-label="">
            </div>  
    '''

    bhtml = '''
            <div class="blank" style="background: #eeeeee; height: 1px; margin: 30px 0 15px;"></div>
            <div style="position: relative; text-align: center">        
    '''

    if data_uuid:
        bhtml += f'''
            <span class="kfConfirmBtn" onclick="edit_site_data('{crr_data.get('uuid')}')">确定</span>                
        '''
    else:
        bhtml += '''
            <span class="kfConfirmBtn" onclick="post_site_data()">确定</span>                
        '''
    bhtml += '''
            <span class="kfCancelBtn" onclick="xtalert.close()">取消</span>
        </div>
    '''
    html = '''
    <div class="site_form_html" style="position: relative; overflow-y: scroll; height: 390px; padding: 20px; text-align: left;">
    %s
    </div>        
    %s
    ''' % (html, bhtml)
    html = update_language(language, html)
    return html


# 清除网站
def del_site_data(site_code):
    BlacklistTable.delete_many({'site_code': site_code})
    CustomerTable.delete_many({'site_code': site_code})
    ChatConversationTable.delete_many({'site_code': site_code})
    LeavingMessageTable.delete_many({'site_code': site_code})
    QuickReplyTable.delete_many({'site_code': site_code})
    problemTable.delete_many({'site_code': site_code})
    systemLogTable.delete_many({'site_code': site_code})
    CmsUserModel.delete_many({'responsible_site': site_code})


# 删除用户
def del_userManage(data_uuid):
    if not data_uuid:
        return xtjson.json_params_error('缺少数据id！')

    _u = CmsUserModel.find_one({'uuid': data_uuid})
    if not _u:
        return xtjson.json_params_error('该用户不存在！')
    _u_role_code = _u.get('role_code')
    if _u_role_code == PermissionCls.AgentAdmin:
        if CmsUserModel.count({'super_admin_id': data_uuid}):
            return xtjson.json_params_error('改代理下存在接入网站，不可删除！')
        CmsUserModel.delete_one({'uuid': data_uuid})
    elif _u_role_code == PermissionCls.SUPERADMIN:
        return xtjson.json_params_error('[系统]管理员用户，不可删除！')
    elif _u_role_code == PermissionCls.Administrator:
        del_site_data(_u)
        CmsUserModel.delete_one({'uuid': data_uuid})
    else:
        CmsUserModel.delete_one({'uuid': data_uuid})

    condatas = ChatConversationTable.find_many({'service_id': data_uuid}) or []
    for cd in condatas:
        ChatContentTable.delete_many({'conversation_id': cd.get('uuid')})

    QuickReplyTable.delete_many({'service_id': data_uuid})
    FinishListTable.delete_many({'service_id': data_uuid})
    # systemLogTable.delete_many({'user_id': data_uuid})
    # _log = {
    #     'user_id': self.current_admin_dict.get('uuid'),
    #     'note': '删除用户！用户账户',
    #     'ip': get_ip(),
    # }
    # systemLogTable.insert_one(_log)
    return xtjson.json_result()


# 客服下线处理
async def disconnect_func(request_sid):
    for k in list(SERVICE_CONNECTION):
        _ddd = SERVICE_CONNECTION.get(k)
        if not _ddd:
            continue
        if _ddd.get('sid') == str(request_sid):
            _d = {
                'crr_service_id': k,
                'serviceState': '0',
            }
            SERVICE_CONNECTION.pop(k)
            await emit('serviceOnlineState', xtjsonCls.json_result(data=_d), namespace='/serviceChat', broadcast=True)

            _udata = CmsUserModel.find_one({'uuid': k}) or {}
            if _udata:
                _log = {
                    'user_id': k,
                    'operation_type': OPERATION_TYPES.OFFLINE,
                    'site_code': _udata.get('responsible_site') or '',
                    'note': '',
                    'ip': _ddd.get('ip') or '',
                }
                systemLogTable.insert_one(_log)


# 向text添加链接
def add_link_a(text):
    if not text:
        return text
    if '<' in text and '>' in text:
        return text
    try:
        pattern = 'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
        urls = re.findall(pattern, text) or []
        for u in urls:
            html = '<a href="%s" target="_blank">%s</a>' % (u, u)
            text = text.replace(u, html, 1)
        return text
    except:
        return text


def format_time_func(data, formatStr=None):
    try:
        if not isinstance(data, datetime.datetime):
            return data
        if not formatStr:
            return data.strftime('%Y-%m-%d %H:%M:%S')
        return data.strftime(formatStr)
    except:
        return data


# 导出数据
def exportDataLy(datas, log_uuid, export_folder, filename):
    ''' 导出数据 '''
    export_data = ExportDataModel.find_one({'uuid': log_uuid})
    if not export_data:
        return
    try:
        if not os.path.exists(export_folder):
            os.makedirs(export_folder)
        crr_count = 0

        wb = Workbook()
        wa = wb.active
        row = 1
        header = ['渠道', '访客ID', 'IP', '姓名', '电话', '邮箱', '留言内容', '时间']
        for h in range(len(header)):
            wa.cell(row=row, column=h+1, value=header[h])

        for data in datas:
            row += 1
            _crr_site_data = SITE_DICT_CACHE.get(data.get('site_code')) or {}
            customer_data = CustomerTable.find_one({'uuid': data.get('customer_id')}) or {}

            wa.cell(row=row, column=1, value=str(_crr_site_data.get('site_name') or ''))
            wa.cell(row=row, column=2, value=str(customer_data.get('name') or ''))
            wa.cell(row=row, column=3, value=str(data.get('ip') or ''))
            wa.cell(row=row, column=4, value=str(data.get('username') or ''))
            wa.cell(row=row, column=5, value=str(data.get('telephone') or ''))
            wa.cell(row=row, column=6, value=str(data.get('email') or ''))
            wa.cell(row=row, column=7, value=str(data.get('text') or ''))
            wa.cell(row=row, column=8, value=format_time_func(data.get('_create_time')))

            crr_count += 1
            if crr_count % 100 == 0:
                export_data['out_count'] = crr_count
                ExportDataModel.save(export_data)

        file_path = os.path.join(export_folder, filename)
        wb.save(file_path)
        export_data['out_count'] = crr_count
        export_data['statu'] = ExportStatu.successed
        ExportDataModel.save(export_data)
        return True
    except Exception as e:
        export_data['note'] = str(e)
        export_data['statu'] = ExportStatu.failed
        ExportDataModel.save(export_data)
        return


